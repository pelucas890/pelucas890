import string
from openpyxl import load_workbook
from openpyxl.xml.constants import MAX_ROW

FILE_PATH = 'E:\Documentos\inva\INVA2.xlsx'

workbook = load_workbook(FILE_PATH)
wb = workbook.active
busqueda = ''
while busqueda != '6':
    print('')
    print('1 Para sumar asistencias')
    print('2 Para restar asist 34 asist')
    print('3 Para restar 2 box al stock')
    print('4 Para agregar una box al stock')
    print('5 Para guardar y salir')
    print('6 Para salir solamente sin guardar')
    print('7 Para restar 20 asist')
    print('8 Para restar 3 asist semanales')
    print('Programa creado por Lucas Lescano alias Desxz')
    print('')
    busqueda = input('Ingresa la opcion: \n')
    print('')
    if busqueda == '1':
        while busqueda != 'SALIR':
                busqueda = input('Introduce la lista: ')         
                busqueda = busqueda.upper()
                for colA,colNB,colC, colND,colE, colNF,colG, colNH, colI, colNJ in wb:
                    if busqueda in colA.value:
                        wb[f'B{colNB.row}'] = colNB.value + 1
                        print(f"Nombre encontrado: {busqueda}. Asistencias: {wb[f'B{colNB.row}'].value}\n")
                    elif busqueda in colC.value:
                        wb[f'D{colND.row}'] = colND.value + 1
                        print(f"Nombre encontrado: {busqueda}. Asistencias: {wb[f'D{colND.row}'].value}\n")
                    elif busqueda in colE.value:
                        wb[f'F{colNF.row}'] = colNF.value + 1
                        print(f"Nombre encontrado: {busqueda}. Asistencias: {wb[f'F{colNF.row}'].value}\n")
                    elif busqueda in colG.value:
                        wb[f'H{colNH.row}'] = colNH.value + 1
                        print(f"Nombre encontrado: {busqueda}. Asistencias: {wb[f'H{colNH.row}'].value}\n")
                    elif busqueda in colI.value:
                        wb[f'J{colNJ.row}'] = colNJ.value + 1
                        print(f"Nombre encontrado: {busqueda}. Asistencias: {wb[f'J{colNJ.row}'].value}\n")
    if busqueda == '2':
        busqueda = input('Ingrese el nombre para restar 34 asist: ')
        busqueda = busqueda.upper()
        for colA,colNB,colC, colND,colE, colNF,colG, colNH, colI, colNJ in wb:
            if busqueda in colA.value:
                wb[f'B{colNB.row}'] = colNB.value - 34
                print(f"COBRO: {busqueda}. Asistencias: {wb[f'B{colNB.row}'].value}\n")
            elif busqueda in colC.value:
                wb[f'D{colND.row}'] = colND.value - 34
                print(f"COBRO: {busqueda}. Asistencias: {wb[f'D{colND.row}'].value}\n")
            elif busqueda in colE.value:
                wb[f'F{colNF.row}'] = colNF.value - 34
                print(f"COBRO: {busqueda}. Asistencias: {wb[f'F{colNF.row}'].value}\n")
            elif busqueda in colG.value:
                wb[f'H{colNH.row}'] = colNH.value - 34
                print(f"COBRO: {busqueda}. Asistencias: {wb[f'H{colNH.row}'].value}\n")
            elif busqueda in colI.value:
                wb[f'J{colNJ.row}'] = colNJ.value - 34
                print(f"COBRO: {busqueda}. Asistencias: {wb[f'J{colNJ.row}'].value}\n")
    if busqueda == '7':
        busqueda = input('Ingrese el nombre para restar 20 asist: ')
        busqueda = busqueda.upper()
        for colA,colNB,colC, colND,colE, colNF,colG, colNH, colI, colNJ in wb:
            if busqueda in colA.value:
                wb[f'B{colNB.row}'] = colNB.value - 20
                print(f"COBRO: {busqueda}. Asistencias: {wb[f'B{colNB.row}'].value}\n")
            elif busqueda in colC.value:
                wb[f'D{colND.row}'] = colND.value - 20
                print(f"COBRO: {busqueda}. Asistencias: {wb[f'D{colND.row}'].value}\n")
            elif busqueda in colE.value:
                wb[f'F{colNF.row}'] = colNF.value - 20
                print(f"COBRO: {busqueda}. Asistencias: {wb[f'F{colNF.row}'].value}\n")
            elif busqueda in colG.value:
                wb[f'H{colNH.row}'] = colNH.value - 20
                print(f"COBRO: {busqueda}. Asistencias: {wb[f'H{colNH.row}'].value}\n")
            elif busqueda in colI.value:
                wb[f'J{colNJ.row}'] = colNJ.value - 20
                print(f"COBRO: {busqueda}. Asistencias: {wb[f'J{colNJ.row}'].value}\n")
    if busqueda == '3':           
        busqueda = 'STOCK'
        for colA,colNB,colC, colND,colE, colNF,colG, colNH, colI, colNJ in wb:
            if busqueda in colA.value:
                wb[f'B{colNB.row}'] = colNB.value - 2
                print(f"Se restaron dos box al stock: {wb[f'B{colNB.row}'].value}\n")
                busqueda = ''
    if busqueda == '4':
        busqueda = 'STOCK'
        for colA,colNB,colC, colND,colE, colNF,colG, colNH, colI, colNJ in wb:
            if busqueda in colA.value:
                wb[f'B{colNB.row}'] = colNB.value + 1
                print(f"Se sumo una box al stock: {wb[f'B{colNB.row}'].value}\n")
    if busqueda == '5':       
        workbook.save('E:\Documentos\inva\Inva2.xlsx')
        print('Guardado Exitoso\n'.center(50, "="))
        break
    if busqueda == '6':
        break
    if busqueda == '8':
        print('Restando 3 asist semanales\n')
        for row in range(2, wb.max_row+1):
            cell = wb.cell(row=row, column=2)
            if cell.value is not wb['B17']:
                cell.value = cell.value - 3
                print('Restando en la columna 2')
        for row in range(2, wb.max_row+1):
            cell = wb.cell(row=row, column=4)
            if cell.value is not wb['B17']:
                cell.value = cell.value - 3
                print('Restando en la columna 4')
        for row in range(2, wb.max_row+1):
            cell = wb.cell(row=row, column=6)
            if cell.value is not wb['B17']:
                cell.value = cell.value - 3
                print('Restando en la columna 6')
        for row in range(2, wb.max_row+1):
            cell = wb.cell(row=row, column=8)
            if cell.value is not wb['B17']:
                cell.value = cell.value - 3
                print('Restando en la columna 8')
        for row in range(2, wb.max_row+1):
            cell = wb.cell(row=row, column=10)
            if cell.value is not wb['B17']:
                cell.value = cell.value - 3                
                print('Restando en la columna 10')
        stock = wb['B17']
        stock.value = stock.value + 3
        print(stock.value)